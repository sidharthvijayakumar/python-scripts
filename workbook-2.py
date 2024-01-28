from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Sid.xlsx')
ws = wb.active

#how to name a excel sheet
ws.title= "Data123"

#Append data to excel sheet
#ws.append(['Sidharth','is','a','good','human'])

#add data to multiple rows usin excel sheet
# for row in range(1,11):
#     for col in range(1,6):
#         char= get_column_letter(col)
#         ws[char+str(row)].value=char+str(row)

#code to delete values in many cells
# for row in range(1,18):
#     for col in range(1,6):
#         ws[get_column_letter(col)+str(row)]= None

#code to merge cells
ws.merge_cells("A1:D1")

#code to unmerge cells
ws.unmerge_cells("A1:D1")

#code to insert a row
#ws.insert_rows(2)

#code to remove rows
#ws.delete_rows(3)

#code to insert colums
#ws.insert_cols(1)

#code to remove colums
ws.delete_cols(1)

#shift the cells from one location to other
ws.move_range("C5:f6",rows=0,cols=0)

wb.save('Sid.xlsx')
