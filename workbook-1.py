from openpyxl import Workbook, load_workbook

wb=load_workbook('Book1.xlsx')
#sets current sheet as active sheet
#ws= wb.active

#access the value of A1 cell in active sheet 
#print(ws['A1'].value)  

#change the value of A2 cell in active sheet
#ws['A2'].value= 'Sid' 

#Prints the value of A2 cell which was changed
#print(ws['A2'].value)

#saves the workbook
#wb.save('Book1.xlsx')

print(wb.sheetnames)

#Creates a sheet with name Grade
#wb.create_sheet('Grade')

#deletes a sheet with name sheet

#print(wb.sheetnames)

#This removes the sheet called Test1
wb.remove(wb['Test1'])

wb.save('Book1.xlsx')
