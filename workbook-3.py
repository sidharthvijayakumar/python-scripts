from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb= Workbook()
ws=wb.active
ws.title="Grades"

#prints the heading of the data in excel sheet
headings=['Name'] + list(data['Joe'].keys())
ws.append(headings)

#prints the whole data in the sheet
for person in data:
    grades=list(data[person].values())
    print(grades)
    ws.append([person]+ grades)

#logi  to calculate avg for a person in the class
my_list=[]

for person in data:
    grades=list(data[person].values())
    total_grade=0
    avg=0
    for mark in grades:
        total_grade+=mark
        avg=total_grade/len(grades)
    my_list.append(avg)
    # for el in range(0,len(data)):
    #     ws['F'+str(el+2)]= avg
    #ws['F'+str(count)]= avg 
    ##logic did not work out 


#logic to calculate avg of a subject
    
for cols in range(2,2+len(list(data['Joe'].values()))):
    char=get_column_letter(cols)
    ws[char+"7"]= f"=SUM({char + '2'}:{char + '6' })/{len(data)}"
#styling the font
for col in range(1,6):
    ws[get_column_letter(col)+str(1)].font= Font(bold=True)
    ws[get_column_letter(cols)+str(1)].alignment=Alignment(horizontal='center', vertical='center')
wb.save('Grades.xlsx')